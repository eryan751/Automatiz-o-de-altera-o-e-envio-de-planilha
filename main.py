import pandas as pd
import os
import smtplib
from email.message import EmailMessage
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook

def update_sheet(planilha_original, linha, coluna):
    arquivo = load_workbook(planilha_original)
    aba_utilizada = arquivo.active
    valor_Cell = aba_utilizada.cell(row=linha, column=coluna).value
    novo_valor = valor_Cell * 1.02
    aba_utilizada.cell(row=linha, column=coluna).value = novo_valor
    arquivo.save("Reservar 2.xlsx")

linha = int(input("Qual a linha desejada: "))
coluna = int(input("Qual a coluna desejada: "))


update_sheet("Reservar 2.xlsx",linha=linha, coluna=coluna)


#Lendo e fazendo as alterações necessarias na planilha originial
dados = pd.read_excel("Reservar 2.xlsx", header=2, na_values = ["0"] )
pd.set_option('display.max_columns', None)

dados_utilizados = ["Produtos", "Qtd_vendida", "Valor Total"]
df = dados[dados_utilizados]


#Salvando a nova versão
df.to_excel('dados_finais.xlsx', index=False)


#Definindo o email remetente
EMAIL_ADDRESS = 'emanueler8010@gmail.com'
#Senha gerada para este app em especifico
EMAIL_PASSWORD = 'mjmxfiyjszawxdrr'

#Assunto e email destinatario
msg = MIMEMultipart()
msg['Subject'] = "Fatutamento do Mês"
msg['From'] = 'emanueler8010@gmail.com'
msg['To'] = 'emadevautonext@gmail.com'
msg.attach(EmailMessage("Faturamento deste Mês"))

#Escolhendo e lendo Anexo que será enviao
caminho_arquivo = "dados_finais.xlsx"
with open(caminho_arquivo, 'rb') as arquivo:
    parte = MIMEBase('application', 'octet-stream')
    parte.set_payload(arquivo.read())

encoders.encode_base64(parte)
parte.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(caminho_arquivo)}"')
msg.attach(parte)

#Enviando
with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
    smtp.login(EMAIL_ADDRESS,EMAIL_PASSWORD)
    smtp.send_message(msg)

print("Email Enviado")


